#!/usr/bin/env python
# coding: utf-8

# In[1]:


from tkinter import *
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Color, numbers
import xlsxwriter
import xlwings as xw
import numpy as np
import pandas as pd
import re
import glob


def find_index(data, target):
    res = []
    lis = data
    while True:
        try:
            res.append(lis.index(target) +
                       (res[-1] + 1 if len(res) != 0 else 0))
            lis = data[res[-1] + 1:]
        except:
            break
    return res


def toexcel(i):
    f = open(i)
    리드라인 = f.readlines()  # 성공
    리드라인_스트링 = "".join(리드라인)  # 리스트를 스트링으로 변환
    리드라인_스드링_스트립 = 리드라인_스트링.strip('<cc:Organization.Name>\n',)

    Content = re.compile('(?<=\<Text.Content>)(.*?)(?=<\/Text.Content>)')
    Content_tag_list = Content.findall(리드라인_스드링_스트립)
    지체상금률 = Content_tag_list[-1]

    물품정보 = Content_tag_list[42:]
    del(물품정보[-5:])
    물품정보수 = len(물품정보)
    품명 = []
    규격 = []
    for i in range(0, 물품정보수, 10):
        품명.append(물품정보[i])
    for i in range(2, 물품정보수, 10):
        규격.append(물품정보[i])

    c = re.compile('(?<=\<Numeric.Content>)(.*?)(?=<\/Numeric.Content>)')
    c_tag_list = c.findall(리드라인_스드링_스트립)

    b = re.compile('(?<=\<Identifier.Content>)(.*?)(?=<\/Identifier.Content>)')
    b_tag_list = b.findall(리드라인_스드링_스트립)
    기관정보 = b.findall(리드라인_스드링_스트립)[:22]
    총목록번호 = b.findall(리드라인_스드링_스트립)[22:]
    물품번호수 = len(총목록번호)
    물품분류번호 = []
    물품식별번호 = []
    for i in range(0, 물품번호수, 5):
        물품분류번호.append(총목록번호[i])
    for i in range(1, 물품번호수, 5):
        물품식별번호.append(총목록번호[i])

    전체수량추출 = re.compile(
        '(?<=\<Quantity.Content>)(.*?)(?=<\/Quantity.Content>)')
    전체수량_tag_list = 전체수량추출.findall(리드라인_스드링_스트립)
    전체수량수 = len(전체수량_tag_list)
    수량 = []
    for i in range(0, 전체수량수, 7):
        수량.append(전체수량_tag_list[i])
    del(수량[-1])

    e = re.compile('(?<=\<Quantity.Unit.Code>)(.*?)(?=<\/Quantity.Unit.Code>)')
    e_tag_list = e.findall(리드라인_스드링_스트립)
    단위 = []
    for i in range(0, len(e_tag_list), 7):
        단위.append(e_tag_list[i])
    del(단위[-1:])

    a = re.compile('(?<=\<Amount.Content>)(.*?)(?=<\/Amount.Content>)')
    총가격_tag_list = a.findall(리드라인_스드링_스트립)
    금액수 = len(총가격_tag_list)
    계약금액일람 = 총가격_tag_list[-8:]
    계약금액 = 계약금액일람[0]
    수수료 = 계약금액일람[-5]
    총납부금액 = 계약금액일람[-4]
    del(총가격_tag_list[-11:])
    가격수 = len(총가격_tag_list)
    단위가격 = []
    합계가격 = []
    for i in range(3, 가격수, 5):
        단위가격.append(총가격_tag_list[i])
    for i in range(4, 가격수, 5):
        합계가격.append(총가격_tag_list[i])

    f = re.compile('(?<=\<DateTime.Content>)(.*?)(?=<\/DateTime.Content>)')
    총날짜_tag_list = f.findall(리드라인_스드링_스트립)
    총계약일_tag_list = f.findall(리드라인_스드링_스트립)[:5]
    총납기_tag_list = f.findall(리드라인_스드링_스트립)[5:]
    납기수 = len(총납기_tag_list)
    납기 = []
    for i in range(0, 납기수, 3):
        납기.append(총납기_tag_list[i])

    # 조건 코드
    g = re.compile('(?<=\<Code.Name>)(.*?)(?=<\/Code.Name>)')
    납품조건_tag_list = g.findall(리드라인_스드링_스트립)[4:]
    납품조건 = []
    총납품조건수 = len(납품조건_tag_list)
    for i in range(1, 총납품조건수, 4):
        납품조건.append(납품조건_tag_list[i])

    wb = Workbook()  # Workbook 생성
    sm_wb = wb.create_sheet('List')  # List 시트 생성
    wb.remove(wb['Sheet'])  # 빈 시트 삭제

    style_center = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
    title_font = Font(size=20, bold=True, color='ffffff')
    subject_font = Font(size=10, bold=True, color='ffffff')
    subject_bg = PatternFill('solid', fgColor='4BACC6')  # 채우기

    border = Border(left=Side(border_style='thin', color='368195'),  # 테두리
                    right=Side(border_style='thin', color='368195'),
                    top=Side(border_style='thin', color='368195'),
                    bottom=Side(border_style='thin', color='368195'))

    subject_range = ['B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'i10', 'j10',
                     'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9',
                     'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9']
    for pos in subject_range:
        sm_wb[pos].font = subject_font
        sm_wb[pos].fill = subject_bg
        sm_wb[pos].alignment = style_center  # 가운데 정렬
        sm_wb[pos].border = border

    sm_wb.freeze_panes = 'A11'  # 6행에 틀고정 설정

    sm_wb.row_dimensions[1].height = 8  # 셀넓이
    sm_wb.row_dimensions[2].height = 40
    sm_wb.row_dimensions[3].height = 16
    sm_wb.column_dimensions['A'].width = 1
    sm_wb.column_dimensions['B'].width = 20
    sm_wb.column_dimensions['C'].width = 65
    sm_wb.column_dimensions['D'].width = 5
    sm_wb.column_dimensions['E'].width = 20
    sm_wb.column_dimensions['F'].width = 15
    sm_wb.column_dimensions['G'].width = 20  # 납품조건
    sm_wb.column_dimensions['H'].width = 10

    sm_wb['B2'].value = '분할납품요구서'
    sm_wb['B2'].font = title_font
    sm_wb['B2'].alignment = style_center
    sm_wb['B2'].fill = PatternFill('solid', fgColor='4BACC6')
    sm_wb['B2'].border = border

    sm_wb.merge_cells('B2:j2')  # 셀병합
    sm_wb.merge_cells('C10:D10')

    '''sm_wb['H4'].value = '공고일 : '   # 공고일은 확인하는 날짜를 기록
    sm_wb['H4'].font = Font(size=10, color='404040')
    sm_wb['H4'].alignment =  Alignment(horizontal='right',vertical='center')'''

    # 왼쪽 인덱스
    sm_wb['B3'].value = '납품요구번호'
    sm_wb['B4'].value = '요청번호'
    sm_wb['B5'].value = '요청건명'
    sm_wb['B6'].value = '납품요구일자'
    sm_wb['B7'].value = '품대계'
    sm_wb['B8'].value = '수수료'
    sm_wb['B9'].value = '합계'

    # 왼쪽 값
    sm_wb['C3'].value = Content_tag_list[2]  # '납품요구번호'
    sm_wb['C4'].value = Content_tag_list[23]  # '요청번호'
    sm_wb['C5'].value = Content_tag_list[28]  # '요청건명'
    sm_wb['C6'].value = 총계약일_tag_list[0]  # '납품요구일자'
    sm_wb['C7'].value = 계약금액  # '품대계'
    sm_wb['C8'].value = 수수료  # '수수료'
    sm_wb['C9'].value = 총납부금액  # '합계'

    # 오른쪽 인덱스
    sm_wb['E3'].value = '계약자'
    sm_wb['E4'].value = '대표자명'
    sm_wb['E5'].value = '담당전화'
    sm_wb['E6'].value = '담당팩스'
    sm_wb['E7'].value = '업체주소'
    sm_wb['E8'].value = ''
    sm_wb['E9'].value = ''

    # 오른쪽 값
    sm_wb['F3'].value = Content_tag_list[14]  # '계약자'
    sm_wb['F4'].value = Content_tag_list[19]  # '대표자명'
    sm_wb['F5'].value = Content_tag_list[20]  # '담당전화'
    sm_wb['F6'].value = Content_tag_list[21]  # '담당팩스'
    sm_wb['F7'].value = Content_tag_list[16]  # '업체주소'
    # sm_wb['F8'].value = Content_tag_list[14]#'할인율'
    # sm_wb['F9'].value = Content_tag_list[14]#'할인금액'

    # 품목 로네일
    sm_wb['B10'].value = '품명'
    sm_wb['C10'].value = '규격'
    # sm_wb['D10'].value =  #단위
    sm_wb['E10'].value = '물품분류번호 물품식별번호'
    sm_wb['F10'].value = '납품기한(원)'
    sm_wb['G10'].value = '납품조건'
    sm_wb['H10'].value = '수량'
    sm_wb['i10'].value = '단가'
    sm_wb['j10'].value = '금액'

    # sm_wb.cell(11,11,'')

    # 품목 갑
    for i in range(len(품명)):
        sm_wb.cell(11 + i, 2, 품명[i])

    for i in range(len(규격)):
        sm_wb.cell(11 + i, 3, 규격[i])

    for i in range(len(단위)):
        sm_wb.cell(11 + i, 4, 단위[i])

    for i in range(len(물품분류번호)):
        sm_wb.cell(11 + i, 5, 물품분류번호[i] + ' ' + 물품식별번호[i])

    for i in range(len(납기)):
        sm_wb.cell(11 + i, 6, 납기[i])

    for i in range(len(납품조건)):
        sm_wb.cell(11 + i, 7, 납품조건[i])

    for i in range(len(수량)):  # 수량
        sm_wb.cell(11 + i, 8, 수량[i])

    for i in range(len(단위가격)):  # 단가
        sm_wb.cell(11 + i, 9, 단위가격[i])

    for i in range(len(합계가격)):  # 금액
        sm_wb.cell(11 + i, 10, 합계가격[i])

    save_name = '{0}_{1}({2}).xlsx'.format(Content_tag_list[23],
                                           Content_tag_list[28], Content_tag_list[2])
    # ,Content_tag_list[40]) #첨부파일명 생성
    path = 'C:/Users/user/Desktop/스캔/1. 계약서 및 민원처리 내역/{0}/'.format(날짜)
    wb.save(path + save_name)
    df = pd.read_excel(path + save_name)
    df


def 분할납품요구서엑셀변환(i):
    날짜 = input()
    XML리스트 = glob.glob(
        'C:/Users/user/Desktop/스캔/1. 계약서 및 민원처리 내역/{0}/분할*/분할*'.format(날짜))
    # 리스트 내용 치환
    XML리스트_치환 = []
    for i in XML리스트:
        temp = i.replace('\\분할납품요구서', '/분할납품요구서')
        XML리스트_치환.append(temp)
    for i in XML리스트_치환:
        toexcel(i)


# 실행버튼 설정
def pressed():
    label.configure(text="날짜 입력")
    분할납품요구서엑셀변환(i)


def 분할납품요구서엑셀변환(i):
    XML리스트 = glob.glob(
        'C:/Users/user/Desktop/스캔/1. 계약서 및 민원처리 내역/{0}/분할*/분할*'.format(날짜))
    # 리스트 내용 치환
    XML리스트_치환 = []
    for i in XML리스트:
        temp = i.replace('\\분할납품요구서', '/분할납품요구서')
        XML리스트_치환.append(temp)

    for i in XML리스트_치환:
        toexcel(i)


# 확인버튼 설정
def confirm():
    in_text = "입력 내용 : " + input_text.get()
    label.configure(text=in_text)
    날짜 = input_text.get()


# 창정보
window = Tk()
window.title("분할납품요구서 엑셀변환기(by jachky)")
window.geometry('320x240')

# 기본 텍스트 표시
label = Label(window, text="★실행방법 : 날짜입력→변환실행", font=("돋음", 10))
label.grid(column=0, row=0)

label2 = Label(
    window, text="※ 파일경로 : C:/Users/user/Desktop/스캔/1. 계약서 및 민원처리 내역/날짜입력", font=("돋음", 6))
label2.grid(column=0, row=20)


# 실행 버튼
button = Button(window, text="변환실행", bg="blue", fg="white", command=pressed)
button.grid(column=0, row=1)

# 입력 내용 출력창
input_text = Entry(window, width=45)
input_text.grid(column=0, row=2)

# 확인 버튼
button = Button(window, text="날짜입력(ex:20211232)", command=confirm)
button.grid(column=0, row=4)

window.mainloop()


# %%
